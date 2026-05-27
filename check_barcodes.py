#!/usr/bin/env python3
"""
check_barcodes.py
-----------------
Validates BeadChip barcodes in a sample sheet CSV and produces a formatted
Excel report plus a clean CSV of all samples.

Usage:
    python check_barcodes.py <input_csv>
    ./check_barcodes.py <input_csv>          # if executable

    If no argument is given, defaults to 'made_samplesheet.csv' in the
    current working directory.

Input:
    A CSV file with at minimum the following columns:
        - Sample ID
        - BeadChip Barcode
        - Plate Number
        - Well Position
        - Collected Gender

Outputs:
    barcode_report.xlsx     -- Excel workbook with three sheets:
                                 1. Summary        : high-level counts
                                 2. Invalid Barcodes: any barcodes not exactly 12 digits
                                 3. All Samples    : full sample list across all plates
    all_plates_samples.csv  -- All samples from the input, written back out as CSV

Dependencies:
    openpyxl  (pip install openpyxl)
"""

import csv
import sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

def style_header(cell, bg_color="4472C4"):
    """Apply bold white text on a coloured background to a header cell."""
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    cell.fill = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def style_cell(cell):
    """Apply standard body text formatting to a data cell."""
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def thin_border():
    """Return a thin light-grey border for all four sides of a cell."""
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


# ---------------------------------------------------------------------------
# Main logic
# ---------------------------------------------------------------------------

def check_barcodes(filepath):
    """
    Read the sample sheet CSV, validate barcodes, and write the Excel report
    and CSV outputs.

    Args:
        filepath (str): Path to the input CSV file.
    """

    all_rows        = []          # Every row from the CSV
    invalid_barcodes = []         # Rows where the barcode length != 12
    barcode_counts  = defaultdict(int)  # Occurrence count per barcode (detects duplicates)
    plates_seen     = []          # Ordered list of unique plate names encountered
    plate_set       = set()       # Fast-lookup set to avoid duplicates in plates_seen

    # --- Parse input CSV ---
    with open(filepath, newline='') as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames  # Preserve original column order for output

        for row in reader:
            all_rows.append(row)

            barcode = row['BeadChip Barcode'].strip()
            barcode_counts[barcode] += 1

            # Flag any barcode that is not exactly 12 characters long
            if len(barcode) != 12:
                invalid_barcodes.append({
                    'Sample ID':       row['Sample ID'].strip(),
                    'BeadChip Barcode': barcode,
                    'Length':          len(barcode),
                    'Plate Number':    row['Plate Number'].strip(),
                    'Well Position':   row['Well Position'].strip(),
                    'Collected Gender': row['Collected Gender'].strip()
                })

            # Track unique plates in the order they first appear
            plate = row['Plate Number'].strip()
            if plate not in plate_set:
                plate_set.add(plate)
                plates_seen.append(plate)

    # --- Write all samples back out as a CSV ---
    csv_out = "all_plates_samples.csv"
    with open(csv_out, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    # --- Build Excel workbook ---
    wb = Workbook()

    # -- Sheet 1: Summary --
    # A quick-glance overview of the key counts from the run.
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_data = [
        ("Total samples processed",       len(all_rows)),
        ("Total plates found",            len(plates_seen)),
        ("Unique barcodes found",         len(barcode_counts)),
        ("Invalid barcodes (<> 12 digits)", len(invalid_barcodes)),
    ]

    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 15

    for i, (label, value) in enumerate(summary_data, start=1):
        cell_a = ws_summary.cell(row=i, column=1, value=label)
        cell_b = ws_summary.cell(row=i, column=2, value=value)
        style_cell(cell_a)
        style_cell(cell_b)
        cell_b.alignment = Alignment(horizontal="center")
        cell_a.border = thin_border()
        cell_b.border = thin_border()
        # Alternate row shading for readability
        if i % 2 == 0:
            for c in [cell_a, cell_b]:
                c.fill = PatternFill("solid", start_color="EEF2FF")

    # -- Sheet 2: Invalid Barcodes --
    # Lists every sample whose BeadChip barcode is not exactly 12 digits.
    # If all barcodes pass validation, a single confirmation message is shown instead.
    ws_invalid = wb.create_sheet("Invalid Barcodes")
    inv_headers = ["Sample ID", "BeadChip Barcode", "Length", "Plate Number", "Well Position", "Collected Gender"]
    col_widths  = [15, 20, 10, 15, 15, 18]

    for col, (header, width) in enumerate(zip(inv_headers, col_widths), start=1):
        cell = ws_invalid.cell(row=1, column=col, value=header)
        style_header(cell)
        ws_invalid.column_dimensions[get_column_letter(col)].width = width

    if invalid_barcodes:
        for row_idx, entry in enumerate(invalid_barcodes, start=2):
            for col_idx, key in enumerate(inv_headers, start=1):
                cell = ws_invalid.cell(row=row_idx, column=col_idx, value=entry[key])
                style_cell(cell)
                cell.border = thin_border()
                if row_idx % 2 == 0:
                    cell.fill = PatternFill("solid", start_color="FFF2CC")
    else:
        # No invalid barcodes found — write a green confirmation message
        ws_invalid.cell(row=2, column=1, value="All barcodes are exactly 12 digits ✓").font = Font(
            name="Arial", size=10, color="375623", bold=True
        )

    # -- Sheet 3: All Samples --
    # Full sample list with all original columns, across every plate in the file.
    # Column widths default to 18; adjust below if needed for wider fields.
    ws_plates = wb.create_sheet("All Samples")
    default_col_width = 18

    for col, header in enumerate(fieldnames, start=1):
        cell = ws_plates.cell(row=1, column=col, value=header)
        style_header(cell, bg_color="375623")
        ws_plates.column_dimensions[get_column_letter(col)].width = default_col_width

    for row_idx, row in enumerate(all_rows, start=2):
        for col_idx, key in enumerate(fieldnames, start=1):
            cell = ws_plates.cell(row=row_idx, column=col_idx, value=row[key].strip())
            style_cell(cell)
            cell.border = thin_border()
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", start_color="E2EFDA")

    wb.save("barcode_report.xlsx")

    print(f"Done.")
    print(f"  XLSX report  -> barcode_report.xlsx")
    print(f"  CSV output   -> {csv_out}  ({len(all_rows)} samples across {len(plates_seen)} plates)")
    print(f"  Invalid barcodes found: {len(invalid_barcodes)}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    # Accept the CSV path as a command-line argument, or fall back to default
    filepath = sys.argv[1] if len(sys.argv) > 1 else "made_samplesheet.csv"
    check_barcodes(filepath)